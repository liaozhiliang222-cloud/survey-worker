"""用真实问卷交叉表（京东常温牛奶-Output-0320.xlsx）测试 pptx_report 模块。

本脚本演示「数据与渲染分离」的真实落地：
  1. :func:`parse_crosstab` 把腾讯问卷 / SPSS 风格的交叉表导出解析为结构化题目；
  2. 用解析结果组装 :class:`~pptx_report.model.ReportSpec`（纯数据）；
  3. 交给 :class:`~pptx_report.renderer.ReportRenderer` 渲染成 .pptx。

刻意覆盖全部 8 种图表类型与 4 种布局，以检验实现能力。

运行：``PYTHONPATH=D:/调研工具 python -m pptx_report.build_jd_report``
"""

from __future__ import annotations

import os
import re

import openpyxl

from .model import (
    AppendixContent,
    ChartPageContent,
    ChartSpec,
    ChartType,
    CoverContent,
    ExecutiveSummaryContent,
    KPI,
    LayoutType,
    ReportSpec,
    TableData,
    TocContent,
)
from .renderer import ReportRenderer
from .theme import Theme

# 六大人群（交叉表表头顺序）
SEGMENT_ORDER = ["都市中产", "都市蓝领", "都市家庭", "都市Z世代", "小镇中年", "小镇青年"]

SHEET_NAME = "Table (%)"

# 模块缓存：最近一次 parse_crosstab 的维度分组结果
_cached_dimension_groups: list[dict] = []
def _detect_dimension_groups(
    rows: list, header_row_idx: int
) -> list[dict]:
    """从多级表头中检测维度分组（如「整体」「是否购买」「手机价格」等）。

    在数据表头行（header_row_idx）上方查找「分组标签行」：
    该行满足 ``row[0] 为空`` 且非空列数**显著少于**数据表头行，
    每个非空单元格代表一个分组的名称及其起始列位置。
    返回 ``[{"name":"整体", "segments":["总体","荣耀",...]}, ...]`` 。
    """
    if header_row_idx < 2:
        return []

    # 取数据表头行作为段名参考
    hdr = rows[header_row_idx]
    hdr_nonempty = sum(1 for i in range(1, len(hdr)) if hdr[i] is not None)

    all_seg_names = []
    for i in range(1, len(hdr)):
        if hdr[i] is not None:
            n = _norm(hdr[i])
            if n and n != "人群":
                all_seg_names.append((i, n))

    if not all_seg_names:
        return []

    # 向上搜索分组标签行：row[0] 为空，且非空列数明显少于数据表头（< 60%）
    # 这样可以区分「分组层」（3~5 个大组）与「数据表头层」（7+ 个细分段）
    group_label_rows = []
    for ri in range(max(0, header_row_idx - 15), header_row_idx):
        r = rows[ri]
        if r[0] is None:
            ne = sum(
                1 for ci in range(1, min(len(r), len(hdr)))
                if r[ci] is not None and _norm(r[ci]) not in ("", "人群")
            )
            # 分组标签行的非空列数应该远少于数据表头
            if 2 <= ne < hdr_nonempty * 0.7:
                labels = [
                    (ci, _norm(r[ci]))
                    for ci in range(1, min(len(r), len(hdr)))
                    if r[ci] is not None and _norm(r[ci]) not in ("", "人群")
                ]
                group_label_rows.append((ri, labels))

    if not group_label_rows:
        # 无多级表头 → 单组退化
        return [{"name": "全部维度",
                 "segments": [n for _, n in all_seg_names],
                 "cols": [i for i, _ in all_seg_names]}]

    # 使用离表头最近的那个分组标签行（最精确的层级）
    best_ri, best_labels = group_label_rows[-1]

    groups = []
    for gi, (col_start, gname) in enumerate(best_labels):
        next_col = best_labels[gi + 1][0] if gi + 1 < len(best_labels) else len(hdr)
        group_segs = []
        group_cols = []
        for sci, sname in all_seg_names:
            if sci >= next_col:
                break
            if sci >= col_start:
                if sname not in group_segs:
                    group_segs.append(sname)
                group_cols.append(sci)
        if group_segs:
            groups.append({"name": gname, "segments": group_segs, "cols": group_cols})

    if not groups:
        groups = [{"name": "全部维度",
                   "segments": [n for _, n in all_seg_names],
                   "cols": [i for i, _ in all_seg_names]}]

    return groups

    if not groups:
        groups = [{"name": "全部维度", "segments": [n for _, n in all_seg_names]}]

    return groups


def parse_crosstab(path: str, sheet_name: str = None) -> list:
    """解析问卷交叉表导出为题目列表。

    每个题目字典结构::

        {
            "code": "VAR1",
            "title": "省份",
            "categories": ["上海", "北京", ...],       # 选项标签（已清洗）
            "segments": ["Total", "都市中产", ...],    # 列：Total + 各人群
            "data": {"Total": [0.049, ...], "都市中产": [...], ...},
            "stats": {"MEAN": {"Total": 34.1, ...}},   # (MEAN)/(AVG.MENTION) 等
            "base": {"Total": 1680, "都市中产": 473, ...},
        }

    说明：
        - 百分比以小数存储（0.049 ≈ 4.9%），渲染时统一 ×100；
        - ``BASE`` 行与 ``(MEAN)`` 等统计行单独提取，不混入分类数据；
        - 整行均为空的选项（如「其他」无数据）会被跳过。

    通用化（v11 集成 PWA）：
        - ``sheet_name`` 为 None 时自动查找名称含 ``table`` 的工作表，否则取第一个；
        - 人群列不再硬编码 ``SEGMENT_ORDER``，表头行 ``row[1]`` 为
          ``Total/合计/总计/总体`` 时，``row[2:]`` 全部作为人群列自动收集，
          支持任意数量、任意命名的人群维度（品牌、城市、年龄等）。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # 自动选择含百分比数据的 sheet（优先 (%)，其次 table，最后取第一个非目录 sheet）
        candidates = [s for s in wb.sheetnames
                      if "(%)" in s.lower() or "%" in s.lower()]
        if not candidates:
            candidates = [s for s in wb.sheetnames if "table" in s.lower()]
        if not candidates:
            # 排除常见的目录/索引类 sheet
            candidates = [s for s in wb.sheetnames
                          if not any(k in s.lower() for k in ("目录", "index", "toc"))]
        ws = wb[candidates[0]] if candidates else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    questions: list = []
    cur = None
    cur_segments = None

    # 模块缓存：最近一次 parse_crosstab 检测到的维度分组（供 CLI 读取）
    global _cached_dimension_groups
    _cached_dimension_groups = []

    # 表头第 1 列（index=1）的合法"总计/TOTAL"关键词集合
    TOTAL_ALIASES = ("Total", "合计", "总计", "总体", "整体")

    for ri, row in enumerate(rows):
        # 表头行：row[0] 为空，第 2 列为 Total/合计/总计/总体/整体，第 3 列起为人群名。
        # 必须在「row[0] 为 None 即跳过」之前检测，否则表头会被漏掉。
        # 通用化：人群列名不再硬编码，row[2:] 全部自动收集（跳过空值）。
        # 排除"人群"分组占位行（腾讯问卷导出常把 row[2] 写成合并单元格的
        # 占位标签「人群」，真正的列名在紧随其后的第二行表头）。
        # 注意：多级表头文件（如荣耀电商）可能连续 2~3 行都满足此条件，
        # 此时取**最后一行匹配**（列最细的那行）作为正式表头。
        if (
            cur is not None
            and len(row) > 2
            and _norm(row[1]) in TOTAL_ALIASES
            and _norm(row[2]) not in ("", None, "人群")
            and sum(
                1 for i in range(2, len(row))
                if row[i] is not None and _norm(row[i]) != ""
            ) >= 1
        ):
            # 收集表头：从 row[1] 起，遇到**重复段名**即停止。
            # 荣耀电商等导出会在同一行内把"7 品牌"结构在多个分析维度上
            # 重复铺开（整体×7、是否购买×7、…），若全部收集会导致
            # segments 重复 20+ 次、图表错乱。只取第一组有效维度。
            segs = []
            for i in range(1, len(row)):
                if row[i] is None:
                    continue
                name = _norm(row[i])
                if not name or name in ("人群",):
                    continue
                if name in segs:  # 段名重复，说明下一组维度开始
                    break
                segs.append(name)
            if len(segs) < 2:  # 退化情况（只有一个段），退回原逻辑
                segs = [_norm(row[1])] + [
                    _norm(row[i])
                    for i in range(2, len(row))
                    if i < len(row) and row[i] is not None and _norm(row[i]) != ""
                ]
            cur["segments"] = segs
            cur_segments = segs
            # 完整列映射（含所有重复铺开的维度组）：(段名, 列索引)。
            # 取**最后一行匹配**（列最细）作为正式表头，覆盖前面较粗的行。
            seg_cols = [
                (_norm(row[i]), i)
                for i in range(1, len(row))
                if row[i] is not None and _norm(row[i]) not in ("", "人群")
            ]
            cur["seg_cols"] = seg_cols
            # 数据容器：data 按段名（首列出现）键控，data_by_col 按列索引键控（支持维度切换）
            cur["data"] = {s: [] for s in segs}
            cur["data_by_col"] = {col: [] for (_, col) in seg_cols}
            cur["base"] = {}
            cur["base_by_col"] = {}
            cur["stats"] = {}
            cur["stats_by_col"] = {}
            # 每次更新表头都重新检测分组（保留段数最多的那次结果）
            new_groups = _detect_dimension_groups(rows, ri)
            if len(new_groups) > len(_cached_dimension_groups):
                _cached_dimension_groups = new_groups
            # 不 continue —— 多级表头文件中后续行可能也是表头，
            # 取最后匹配的一行（列最细）作为正式表头。

        if row[0] is None:
            continue
        a = _norm(row[0])

        # 章节标记 PART:[...] —— 仅作分隔，不形成题目
        if a.startswith("PART:["):
            continue

        # 题目标记 CAPTION:[VARxx].题面
        if a.startswith("CAPTION:["):
            m = re.match(r"CAPTION:\[([^\]]+)\]\.?\s*(.*)", a)
            code = m.group(1) if m else a
            title = (m.group(2).strip() if m else a)
            cur = {
                "code": code,
                "title": title,
                "categories": [],
                "segments": None,
                "data": {},
                "stats": {},
                "base": {},
            }
            questions.append(cur)
            cur_segments = None
            continue

        if cur is None:
            continue

        # 表头检测已前置到循环顶部（处理 row[0] 为 None 的表头行）
        if cur_segments is None:
            continue

        cat = a

        # 统计行：(MEAN) / (STD.DEV.) / (AVG.MENTION)
        if cat.startswith("("):
            stat = cat.strip("()").split()[0].replace(".", "_")
            for (name, col) in cur["seg_cols"]:
                v = row[col] if col < len(row) else None
                if v is not None:
                    try:
                        cur["stats_by_col"].setdefault(col, {})[stat] = float(v)
                    except (TypeError, ValueError):
                        pass
            # 首列出现 → stats（向后兼容）
            seen = set()
            for (name, col) in cur["seg_cols"]:
                if name in seen:
                    continue
                seen.add(name)
                cur["stats"].setdefault(stat, {})[name] = cur["stats_by_col"].get(col, {}).get(stat)
            continue

        # 样本量行：BASE
        if cat.upper() == "BASE":
            for (name, col) in cur["seg_cols"]:
                v = row[col] if col < len(row) else None
                if v is not None:
                    try:
                        cur["base_by_col"][col] = int(float(v))
                    except (TypeError, ValueError):
                        pass
            # 首列出现 → base（向后兼容）
            seen = set()
            for (name, col) in cur["seg_cols"]:
                if name in seen:
                    continue
                seen.add(name)
                cur["base"][name] = cur["base_by_col"].get(col)
            continue

        # 普通分类数据行：清洗尾部 "（...）" 注释
        clean = re.sub(r"\s*[（(][^（）()]*[）)]\s*$", "", cat).strip()
        col_vals = {}
        all_none = True
        for (name, col) in cur["seg_cols"]:
            v = row[col] if col < len(row) else None
            if v is not None:
                try:
                    fv = float(v)
                    col_vals[col] = fv
                    all_none = False
                except (TypeError, ValueError):
                    pass
        if all_none:
            continue
        cur["categories"].append(clean)
        for (name, col) in cur["seg_cols"]:
            cur["data_by_col"][col].append(col_vals.get(col))
        # 首列出现 → data（向后兼容，默认整体维度）
        seen = set()
        for (name, col) in cur["seg_cols"]:
            if name in seen:
                continue
            seen.add(name)
            cur["data"][name].append(col_vals.get(col))

    return questions


def apply_dimension(questions: list, dimension_groups: list, group_name: str = None) -> list:
    """按选中的「维度分组」重映射每题的 segments / data / base / stats。

    多级表头文件（如荣耀电商）中，各分组列名相同（整体/购买/未购买 下的
    「荣耀」都是同一名字），必须用**列索引**区分。``dimension_groups`` 中
    每组带 ``cols``（列索引列表）。

    Args:
        questions: 解析出的题目列表。
        dimension_groups: 维度分组列表。
        group_name: 分组名，支持以下形式：
            - ``None`` → 取第一个分组（默认行为）
            - 单个名称（如 ``"购买"``）→ 匹配该分组
            - 逗号分隔的多名（如 ``"省份,年龄,性别"``）→ **合并**多个分组的列

    Returns:
        重映射后的题目列表；找不到则原样返回。
    """
    if not dimension_groups:
        return questions

    # 支持逗号分隔的多组选择（前端多选时传 "省份,年龄,性别"）
    targets = [n.strip() for n in (group_name or "").split(",") if n.strip()] if group_name else []

    if len(targets) <= 1:
        # 单选或默认：取第一个命中的组
        grp = None
        if targets:
            for g in dimension_groups:
                if g["name"] == targets[0]:
                    grp = g; break
        if grp is None:
            grp = dimension_groups[0]
        _groups_to_apply = [grp]
    else:
        # 多选：逐个查找，全部命中才生效（忽略未命中的名字）
        _groups_to_apply = []
        for t in targets:
            for g in dimension_groups:
                if g["name"] == t:
                    _groups_to_apply.append(g)
                    break
        if not _groups_to_apply:
            _groups_to_apply = [dimension_groups[0]]

    # 合并所有选中组的 segments + cols（按加入顺序去重保序）
    merged_names = []
    merged_cols = []
    seen_cols = set()
    for g in _groups_to_apply:
        for name, col in zip(g.get("segments", []), g.get("cols", [])):
            if col not in seen_cols:
                seen_cols.add(col)
                merged_names.append(name)
                merged_cols.append(col)

    if not merged_names or not merged_cols:
        return questions

    out = []
    for q in questions:
        by_col = q.get("data_by_col", {})
        base_by_col = q.get("base_by_col", {})
        stats_by_col = q.get("stats_by_col", {})
        nq = dict(q)
        # 所有分维度分析都保留总体列，便于把分群结果与市场整体基准直接比较。
        total_key = next(
            (name for name in (q.get("segments") or []) if str(name).strip().lower() in {"total", "总体", "整体", "合计", "总计"}),
            None,
        )
        selected_names = list(merged_names)
        if total_key and total_key not in selected_names:
            selected_names.insert(0, total_key)
        nq["segments"] = selected_names
        nq["data"] = {name: by_col.get(col, []) for name, col in zip(merged_names, merged_cols)}
        nq["base"] = {name: base_by_col.get(col) for name, col in zip(merged_names, merged_cols)}
        if total_key and total_key not in nq["data"]:
            nq["data"] = {total_key: list((q.get("data") or {}).get(total_key, [])), **nq["data"]}
            nq["base"] = {total_key: (q.get("base") or {}).get(total_key), **nq["base"]}
        new_stats = {}
        for stat in set().union(*[set(sc.keys()) for sc in stats_by_col.values()]) if stats_by_col else []:
            new_stats[stat] = {}
            for name, col in zip(merged_names, merged_cols):
                if col in stats_by_col and stat in stats_by_col[col]:
                    new_stats[stat][name] = stats_by_col[col][stat]
            if total_key:
                total_stat = (q.get("stats") or {}).get(stat, {}).get(total_key)
                if total_stat is not None:
                    new_stats[stat] = {total_key: total_stat, **new_stats[stat]}
        nq["stats"] = new_stats
        out.append(nq)
    return out


def get_q(questions: list, code: str) -> dict:
    """按题号取题目字典。"""
    for q in questions:
        if q["code"] == code:
            return q
    raise KeyError(f"未找到题目 {code}")


def _series_pct(q: dict, seg: str, categories: list) -> list:
    """取某人群在指定分类上的百分比（×100，缺失补 0）。"""
    data = q["data"].get(seg, [])
    out = []
    for i in range(len(categories)):
        v = data[i] if i < len(data) else None
        out.append(round(v * 100, 1) if v is not None else 0.0)
    return out


def _shorten(label: str) -> str:
    """场景等长标签去掉尾部括号注释，便于图表阅读。"""
    return label.split("（")[0].split("(")[0].strip()


def _norm(value) -> str:
    """归一化单元格文本：去首尾空白，并剥掉导出时包裹的单/双引号。

    该交叉表导出把文本单元格包成 ``'Total'`` / ``'都市中产'`` 形式，
    若不处理会导致表头识别与人群列匹配全部失败。
    """
    s = str(value).strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in "\"'":
        s = s[1:-1]
    return s


# ------------------------- 2. 组装报告 -------------------------
def build_report_spec(questions: list) -> ReportSpec:
    """用解析结果组装 ReportSpec（纯数据，与渲染解耦）。"""

    # —— 封面 ——
    cover = CoverContent(
        title="京东常温牛奶消费者洞察报告",
        client="京东",
        date="2026-03-20",
        subtitle="定量调研 · 八大人群 × 消费场景深度分析",
    )

    # —— 执行摘要 KPI ——
    kpis = [
        KPI(label="总样本量", value="1,680"),
        KPI(label="覆盖城市人群", value="6 类"),
        KPI(label="都市蓝领占比", value="33.8%"),
        KPI(label="早餐场景渗透", value="72.9%"),
        KPI(label="周均饮用频次", value="4.7 次"),
    ]
    exec_summary = ExecutiveSummaryContent(
        kpis=kpis,
        conclusion=(
            "常温奶消费以都市人群为基本盘，早餐为核心场景；营养（高蛋白 / 高钙）、"
            "新鲜（保质期短）与性价比是核心购买驱动，家庭人群更关注子女饮用场景。"
        ),
    )

    # —— 解析常用题目 ——
    q_prov = get_q(questions, "VAR1")        # 省份
    q_city = get_q(questions, "VAR11")       # 城市级别
    q_age = get_q(questions, "S2")           # 年龄
    q_edu = get_q(questions, "D1")           # 学历
    q_inc = get_q(questions, "D3")           # 家庭月收入
    q_freq = get_q(questions, "HS8")         # 购买频率
    q_obj = get_q(questions, "HB1")          # 购买对象
    q_scene = get_q(questions, "HB2")        # 消费场景
    q_nutri = get_q(questions, "HA2A")       # 营养成分关注
    q_v1 = get_q(questions, "V1")            # 八大人群构成
    q_hb3 = get_q(questions, "HB3")          # 每周饮用频次（含 MEAN）

    # 图表页集合
    chart_pages: list = []

    # P4 单图大版面：省份分布
    chart_pages.append(
        ChartPageContent(
            title="消费者地域分布",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.bar(
                    title="常温牛奶消费者省份分布（Total，%）",
                    categories=q_prov["categories"],
                    series_dict={"Total": _series_pct(q_prov, "Total", q_prov["categories"])},
                    insight="北京、广东、河北为三大主力省份，合计占比超五成。",
                )
            ],
        )
    )

    # P5 仪表盘 2x2：城市级别 / 年龄 / 学历 / 收入
    chart_pages.append(
        ChartPageContent(
            title="人群基础画像（仪表盘）",
            layout=LayoutType.DASHBOARD,
            charts=[
                ChartSpec.bar(
                    title="城市级别分布（%）",
                    categories=q_city["categories"],
                    series_dict={"Total": _series_pct(q_city, "Total", q_city["categories"])},
                    insight="一线 + 新一线合计约六成。",
                ),
                ChartSpec.line(
                    title="年龄结构（%）",
                    categories=q_age["categories"],
                    series_dict={"Total": _series_pct(q_age, "Total", q_age["categories"])},
                    insight="26-40 岁为主力。",
                ),
                ChartSpec.bar(
                    title="学历分布（%）",
                    categories=q_edu["categories"],
                    series_dict={"Total": _series_pct(q_edu, "Total", q_edu["categories"])},
                    insight="本科及以上过七成。",
                ),
                ChartSpec.bar(
                    title="家庭月收入分布（%）",
                    categories=q_inc["categories"],
                    series_dict={"Total": _series_pct(q_inc, "Total", q_inc["categories"])},
                    insight="8,001-15,000 元为主力收入带。",
                ),
            ],
        )
    )

    # P6 对比式双图：营养成分关注 中产 vs 蓝领
    chart_pages.append(
        ChartPageContent(
            title="营养成分关注度对比（都市中产 vs 都市蓝领）",
            layout=LayoutType.DUAL,
            charts=[
                ChartSpec.bar(
                    title="营养成分关注 — 都市中产（%）",
                    categories=q_nutri["categories"],
                    series_dict={"都市中产": _series_pct(q_nutri, "都市中产", q_nutri["categories"])},
                    insight="中产更关注高蛋白质、高钙、A2-β酪蛋白。",
                ),
                ChartSpec.bar(
                    title="营养成分关注 — 都市蓝领（%）",
                    categories=q_nutri["categories"],
                    series_dict={"都市蓝领": _series_pct(q_nutri, "都市蓝领", q_nutri["categories"])},
                    insight="蓝领同样看重高蛋白，但更在意保质期新鲜。",
                ),
            ],
        )
    )

    # P7 图文混排：消费场景 + 洞察
    scene_cats = q_scene["categories"]
    scene_short = [_shorten(c) for c in scene_cats]
    chart_pages.append(
        ChartPageContent(
            title="消费场景分布",
            layout=LayoutType.MIXED,
            charts=[
                ChartSpec.bar(
                    title="常温牛奶消费场景（Total，%）",
                    categories=scene_short,
                    series_dict={"Total": _series_pct(q_scene, "Total", scene_cats)},
                    insight="早餐为核心场景，居家与睡眠场景为辅。",
                )
            ],
            side_insights=[
                "早餐场景渗透最高（72.9%），是常温奶第一消费场景",
                "睡前饮用（38.5%）、佐餐（31.7%）次之",
                "运动后 / 办公室加餐场景稳步渗透",
                "社交聚会、学习备考场景占比较低",
                "家庭人群在「子女」相关场景显著突出",
            ],
        )
    )

    # P8 单图：八大人群构成（环形图）
    v1_cats = q_v1["categories"]
    chart_pages.append(
        ChartPageContent(
            title="八大人群构成",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.doughnut(
                    title="八大人群占比构成（%）",
                    categories=v1_cats,
                    values=_series_pct(q_v1, "Total", v1_cats),
                    insight="都市蓝领（33.8%）+ 都市中产（28.2%）合计超六成。",
                )
            ],
        )
    )

    # P9 单图组合图：样本量（柱）+ 平均年龄（折线，副轴）
    base = q_prov.get("base") or q_city.get("base") or {}
    mean_age = q_age.get("stats", {}).get("MEAN", {})
    base_list = [base.get(s, 0) for s in SEGMENT_ORDER]
    age_list = [round(mean_age.get(s, 0), 1) for s in SEGMENT_ORDER]
    chart_pages.append(
        ChartPageContent(
            title="人群规模与年龄结构",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.combo(
                    title="各人群样本量（柱）与平均年龄（折线）",
                    categories=SEGMENT_ORDER,
                    bars={"样本量": base_list},
                    line=age_list,
                    line_name="平均年龄",
                    secondary_axis_title="平均年龄",
                    insight="小镇中年样本虽少，但平均年龄最高（近 40 岁）。",
                )
            ],
        )
    )

    # P10 单图雷达：消费场景多维对比（中产 / 蓝领 / 家庭）
    desired = {"早餐时", "佐餐", "闲暇时", "下午茶/DIY 饮品",
               "运动后/健身时补充能量", "睡前饮用助眠", "办公室/工位上", "户外/长途出行时"}
    picks = [(scene_short[i], scene_cats[i])
             for i in range(len(scene_cats)) if scene_short[i] in desired]
    radar_labels = [p[0] for p in picks]
    radar_full = [p[1] for p in picks]
    chart_pages.append(
        ChartPageContent(
            title="消费场景偏好多维对比",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.radar(
                    title="消费场景偏好 — 中产 / 蓝领 / 家庭（%）",
                    categories=radar_labels,
                    series_dict={
                        "都市中产": _series_pct(q_scene, "都市中产", radar_full),
                        "都市蓝领": _series_pct(q_scene, "都市蓝领", radar_full),
                        "都市家庭": _series_pct(q_scene, "都市家庭", radar_full),
                    },
                    insight="家庭人群在「子女」场景外更均衡；中产场景分布最广。",
                )
            ],
        )
    )

    # P11 单图散点：家庭月收入 vs 每周饮用频次
    mean_inc = q_inc.get("stats", {}).get("MEAN", {})
    mean_freq = q_hb3.get("stats", {}).get("MEAN", {})
    x_vals = [round(mean_inc.get(s, 0), 0) for s in SEGMENT_ORDER]
    y_vals = [round(mean_freq.get(s, 0), 2) for s in SEGMENT_ORDER]
    chart_pages.append(
        ChartPageContent(
            title="收入与饮用频次相关性",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.scatter(
                    title="家庭月收入 vs 每周饮用频次（按人群）",
                    x_values=x_vals,
                    y_values=y_vals,
                    name="人群",
                    insight="收入与饮用频次无明显正相关，饮用习惯更受场景驱动。",
                )
            ],
        )
    )

    # P12 单图堆积柱：购买对象构成（按 3 类核心人群）
    obj_cats = q_obj["categories"]
    chart_pages.append(
        ChartPageContent(
            title="购买对象构成",
            layout=LayoutType.SINGLE,
            charts=[
                ChartSpec.bar(
                    title="购买对象构成 — 中产 / 蓝领 / 家庭（%）",
                    categories=obj_cats,
                    series_dict={
                        "都市中产": _series_pct(q_obj, "都市中产", obj_cats),
                        "都市蓝领": _series_pct(q_obj, "都市蓝领", obj_cats),
                        "都市家庭": _series_pct(q_obj, "都市家庭", obj_cats),
                    },
                    insight="均以「自己」饮用为主，「子女」在家庭中显著更高。",
                    stacked=True,
                )
            ],
        )
    )

    # —— 附录表格：人群画像汇总 ——
    mean_inc = q_inc.get("stats", {}).get("MEAN", {})
    mean_freq = q_hb3.get("stats", {}).get("MEAN", {})
    v1_total = q_v1["data"].get("Total", [])
    headers = ["人群", "样本量", "占比(%)", "平均年龄", "月收入均值(元)", "周均饮用(次)"]
    rows = []
    for s in SEGMENT_ORDER:
        i = v1_cats.index(s) if s in v1_cats else None
        share = round(v1_total[i] * 100, 1) if i is not None else 0.0
        rows.append([
            s,
            base.get(s, 0),
            share,
            round(mean_age.get(s, 0), 1),
            round(mean_inc.get(s, 0), 0),
            round(mean_freq.get(s, 0), 2),
        ])
    appendix = AppendixContent(
        title="数据附录 · 人群画像汇总",
        table=TableData(headers=headers, rows=rows),
        source="数据来源：京东常温牛奶消费者调研 Output-0320（N=1,680，6 类城市人群）。",
    )

    return ReportSpec(
        cover=cover,
        toc=TocContent(),
        executive_summary=exec_summary,
        chart_pages=chart_pages,
        appendix=appendix,
    )


# ------------------------- 3. 入口 -------------------------
def main() -> None:
    """生成优化版京东报告：全题目覆盖 + 数据标签 + 排序 + 多维度人群对比。

    通过通用向导 :mod:`pptx_report.wizard` 自动完成，对齐调研公司交付规范。
    （:func:`build_report_spec` 仍保留，作为「手工精选 11 题」版的可读示例。）
    """
    here = os.path.dirname(os.path.abspath(__file__))
    xlsx = os.path.join("C:/Users/a1382/Desktop", "京东常温牛奶-Output-0320.xlsx")
    out_dir = os.path.abspath(os.path.join(here, "..", "outputs"))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "京东常温牛奶_调研报告.pptx")

    from .wizard import run_wizard
    saved = run_wizard(
        xlsx, out_path,
        title="京东常温牛奶消费者洞察报告",
        client="京东",
        date="2026-03-20",
        source="数据来源：京东常温牛奶消费者调研（腾讯问卷，全国样本 N=800）。",
        max_per_page=3,
    )
    print(f"\n报告已生成: {saved}")


if __name__ == "__main__":
    main()
